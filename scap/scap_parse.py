import xml.etree.ElementTree as ET
import openpyxl

def parse_scap_results(input_xml_file, output_excel_file):
    tree = ET.parse(input_xml_file)
    root = tree.getroot()

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ["idref", "result", "time", "severity", "title", "description", "warning", "reference", "rationale", "fix"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col).value = header

    ns = {'xccdf': 'http://checklists.nist.gov/xccdf/1.2'}
    for i, rule in enumerate(root.findall(".//xccdf:rule-result", ns), start=2):
        idref = rule.get("idref")
        xml_rule = root.find(f".//xccdf:Rule[@id='{idref}']", ns)
        if xml_rule is not None:
            values = [
                idref,
                rule.find("xccdf:result", ns).text,
                rule.get("time"),
                rule.get("severity"),
                xml_rule.find("xccdf:title", ns).text,
                xml_rule.find("xccdf:description", ns).text,
                xml_rule.find("xccdf:warning", ns).text if xml_rule.find("xccdf:warning", ns) is not None else '',
                xml_rule.find("xccdf:reference", ns).get("href") if xml_rule.find("xccdf:reference", ns) is not None else '',
                xml_rule.find("xccdf:rationale", ns).text if xml_rule.find("xccdf:rationale", ns) is not None else '',
                xml_rule.find("xccdf:fix", ns).text if xml_rule.find("xccdf:fix", ns) is not None else ''
            ]
            for col, value in enumerate(values, start=1):
                sheet.cell(row=i, column=col).value = value

    workbook.save(output_excel_file)

def main():
    input_xml_file = input("Enter the path to the SCAP results XML file: ")
    output_excel_file = input("Enter the output Excel file name: ")
    parse_scap_results(input_xml_file, output_excel_file)
    print(f"SCAP results have been parsed and saved in {output_excel_file}")

if __name__ == "__main__":
    main()
